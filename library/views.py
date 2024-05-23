from django.http import JsonResponse
from django.shortcuts import redirect, render,HttpResponse
from .models import *
from .forms import UploadFileForm
from .constants import *
from datetime import date, datetime
from django.db.models import Count
from django.core.files.storage import FileSystemStorage
import pandas as pd
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from django.shortcuts import render, redirect

def index(request):
    return render(request, "index.html")

# Error
def error_404(request):
    return render(request, "error_404.html")

def error_500(request):
    return render(request, "error_500.html")

# Book
# @login_required(login_url = '/admin_login')
def view_books(request, categoryId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if categoryId:
            books = Book.objects.filter(category=categoryId)
        else:
            books = Book.objects.all()
        for book in books:
            num = Loan.objects.filter(return_date=None, books__in=[book]).count()
            book.rest = book.quantity - num
        categories = Category.objects.all()
        return render(request, "view_books.html", {'books':books, 'categories':categories})
    except Exception as e:
        print(e)
        return redirect('error_500')

# @login_required(login_url = '/admin_login')
def add_book(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == "POST":
            name = request.POST['name']
            author = request.POST['author']
            category = Category.objects.filter(id=request.POST['category'])[0]
            quantity = request.POST['quantity']
            book = Book.objects.create(name=name, author=author, category=category, quantity=quantity)
            book.save()
            alert = True
            librarian = Account.objects.get(username=request.session['account'])
            Log.objects.create(action=f"Add book with ID = {book.id}",  account=librarian)
            return render(request, "add_book.html", {'alert':alert})
        
        categories = Category.objects.all()
        return render(request, "add_book.html", {'categories':categories})
    except Exception as e:
        print(e)
        return redirect('error_500')

def check_book_id(request, id):
    try:
        data = {
            'is_taken': Book.objects.filter(id=id).exists()
        }
        return JsonResponse(data)
    except Exception as e:
        return redirect('error_500')

def delete_book(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        book = Book.objects.filter(id=id)
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Delete book with ID = {book.id}",  account=librarian)
        book.delete()
        return render(request, 'view_books.html')
    except Exception as e:
        return redirect('error_500')

def edit_book(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        book = Book.objects.get(id=id)
        if request.method == "POST":
            name = request.POST['name']
            author = request.POST['author']
            category = Category.objects.filter(id=request.POST['category'])[0]
            quantity = request.POST['quantity']
            book.name = name
            book.author = author
            book.category = category
            book.quantity = quantity
            book.save()
            librarian = Account.objects.get(username=request.session['account'])
            Log.objects.create(action=f"Update book with ID = {book.id}",  account=librarian)
            alert = True
            return render(request, "edit_book.html", {'alert':alert, 'book':book})
        categories = Category.objects.all()
        return render(request, "edit_book.html", {'book':book, 'categories':categories})
    except Exception as e:
        return redirect('error_500')

# Category
def view_categories(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        categories = Category.objects.annotate(quantity=Count('book'))
        return render(request, "view_categories.html", {'categories': categories})
    except Exception as e:
        return redirect('error_500')

def add_category(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == "POST":
            name = request.POST['name']
            category = Category.objects.create(name=name)
            category.save()
            alert = True
            librarian = Account.objects.get(username=request.session['account'])
            Log.objects.create(action=f"Add category with ID = {category.id}",  account=librarian)
            return render(request, "add_category.html", {'alert':alert})
        return render(request, "add_category.html")
    except Exception as e:
        print(e)
        return redirect('error_500')

def delete_category(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        books = Book.objects.filter(category=id)
        books.delete()
        category = Category.objects.filter(id=id)
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Delete category with ID = {category.id}",  account=librarian)
        category.delete()
        return render(request, "/view_categories")
    except Exception as e:
        return redirect('error_500')

def edit_category(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        category = Category.objects.get(id=id)
        if request.method == "POST":
            name = request.POST['name']
            category.name = name
            librarian = Account.objects.get(username=request.session['account'])
            Log.objects.create(action=f"Update category with ID = {category.id}",  account=librarian)
            category.save()
            alert = True
            return render(request, "edit_category.html", {'alert':alert, 'category':category})
        return render(request, "edit_category.html", {'category':category})
    except Exception as e:
        print(e)
        return redirect('error_500')

def check_category_name(request, name):
    try:
        data = {
            'is_taken': Category.objects.filter(name=name).exists()
        }
        return JsonResponse(data)
    except Exception as e:
        return redirect('error_500')

# Student
# @login_required(login_url = '/admin_login')
def view_students(request, classId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if classId:
            students = Student.objects.filter(studentClass=classId)
        else:
            students = Student.objects.all()
        classes = StudentClass.objects.all()
        return render(request, "view_students.html", {'students':students, 'classes':classes})
    except Exception as e:
        return redirect('error_500')

def check_student_id(request, id):
    try:
        data = {
            'is_taken': Student.objects.filter(id=id).exists()
        }
        return JsonResponse(data)
    except Exception as e:
        return redirect('error_500')

def upload_file_student(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == 'POST' and request.FILES['file']:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                file_path = fs.path(filename)
                
                try:
                    data = pd.read_excel(file_path)
                    for index, row in data.iterrows():
                        account, created = Account.objects.update_or_create(
                            username=row['username'],
                            defaults={
                                'password': row['password'],
                                'role': True,
                            }
                        )
                        student, created = Student.objects.update_or_create(
                            id=row['id'],
                            defaults={
                                'name': row['name'],
                                'studentClass': StudentClass.objects.get(name=row['class']),
                                'account': account
                            }
                        )
                    librarian = Account.objects.get(username=request.session['account'])
                    Log.objects.create(action=f"Import the data about student from the file named '{filename}'",  account=librarian)
                    alert = "File imported successfully"
                except Exception as e:
                    alert = f"Error processing file: {e}"

                return redirect("/view_students")
        else:
            form = UploadFileForm()
        return redirect("/view_students")
    except Exception as e:
        return redirect('error_500')

def download_sample_student(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        columns = ["id", "name", "class", "username", "password"]
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        classes = StudentClass.objects.all()
        class_names = [cls.name for cls in classes]
        class_names_str = ','.join(class_names)

        dv = DataValidation(type="list", formula1=f'"{class_names_str}"', allow_blank=True)
        sheet.add_data_validation(dv)

        for row in range(2, 999):
            dv.add(sheet.cell(row=row, column=3))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample_student.xlsx'

        workbook.save(response)
        return response
    except Exception as e:
        return redirect('error_500')
    
def delete_student(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        student = Student.objects.filter(id=id)
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Delete student with ID = {student.id}",  account=librarian)
        student.delete()
        return render(request, "/view_students")
    except Exception as e:
        return redirect('error_500')

#Loan
def view_loan_history(request, studentId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if studentId:
            loans = Loan.objects.filter(student=Student.objects.get(id=studentId))
        else:
            loans = Loan.objects.all()
        for loan in loans:
            loan.bookIds = ', '.join(str(book.id) for book in loan.books.all())

        return render(request, "view_loan_history.html", {'loanList':loans})
    except Exception as e:
        return redirect('error_500')

def detail_loan(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        loan = Loan.objects.get(id=id)
        books = loan.books.all()
        return render(request, "detail_loan.html", {'loan':loan, 'books':books})
    except Exception as e:
        return redirect('error_500')

def add_loan(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == "POST":
            receivedBookIds = request.POST.get('bookIds')
            bookIds = receivedBookIds.split(',')
            print(bookIds)
            student = Student.objects.get(id=request.POST['studentId'])
            due_date = request.POST['dueDate']
            borrow_date = datetime.now().strftime("%Y-%m-%d")
            loan = Loan.objects.create(student=student, borrow_date=borrow_date, return_date=None, due_date=due_date)
            books_to_add = Book.objects.filter(id__in=bookIds)  # List các id của các Book cần thêm
            loan.books.add(*books_to_add)
            loan.save()
            librarian = Account.objects.get(username=request.session['account'])
            Log.objects.create(action=f"Add a loan with ID = {loan.id}, student ID = {request.POST['studentId']} and book IDs = {receivedBookIds}",  account=librarian)
            alert = True
            return render(request, "add_loan.html", {'alert':alert})
        books = Book.objects.all()
        students = Student.objects.all()
        return render(request, "add_loan.html", {'books':books, 'students':students})
    except Exception as e:
        return redirect('error_500')

def return_loan(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')

        loan = Loan.objects.get(id=id)
        loan.return_date = datetime.now().strftime("%Y-%m-%d")
        loan.save()
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Return a loan with ID = {id}",  account=librarian)
        return redirect('view_loan_history')
    except Exception as e:
        return redirect('error_500')

def renew_loan(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')

        loan = Loan.objects.get(id=id)
        loan.due_date = request.POST['dueDate']
        loan.save()
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Renew a loan with ID = {id}",  account=librarian)
        return redirect('view_loan_history')
    except Exception as e:
        return redirect('error_500')

def delete_loan(request, id):
    try:
        loan = Loan.objects.get(id=id)
        loan.delete()
        librarian = Account.objects.get(username=request.session['account'])
        Log.objects.create(action=f"Delete a loan with ID = {id}",  account=librarian)
        return redirect('view_loan_history')
    except Exception as e:
        return redirect('error_500')

# Faculty
def view_faculties(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        faculties = Faculty.objects.all().annotate(quantity=Count('studentclass'))
        return render(request, "view_faculties.html", {'faculties':faculties})
    except Exception as e:
        return redirect('error_500')

def upload_file_faculty(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == 'POST' and request.FILES['file']:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                file_path = fs.path(filename)
                
                try:
                    data = pd.read_excel(file_path)
                    for index, row in data.iterrows():
                        faculty, created = Faculty.objects.update_or_create(
                            id=row['id'],
                            defaults={'name': row['name']}
                        )
                    librarian = Account.objects.get(username=request.session['account'])
                    Log.objects.create(action=f"Import the data about faculties from the file named '{filename}'",  account=librarian)
                    message = "File imported successfully"
                except Exception as e:
                    message = f"Error processing file: {e}"
                print(message)
                return redirect("/view_faculties")
        else:
            form = UploadFileForm()
        return redirect("/view_faculties")
    except Exception as e:
        return redirect('error_500')

def download_sample_faculty(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        
        columns = ["id", "name"]
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample_faculty.xlsx'

        workbook.save(response)
        return response
    except Exception as e:
        return redirect('error_500')

# Class
def view_classes(request, facultyId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if facultyId:
            classes = StudentClass.objects.filter(faculty=facultyId).annotate(quantity=Count('student'))
        else:
            classes = StudentClass.objects.all().annotate(quantity=Count('student'))
        faculties = Faculty.objects.all()
        return render(request, "view_classes.html", {'classes':classes, 'faculties':faculties})
    except Exception as e:
        return redirect('error_500')

def upload_file_class(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        if request.method == 'POST' and request.FILES['file']:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                file_path = fs.path(filename)
                
                try:
                    data = pd.read_excel(file_path)
                    print(data)
                    for index, row in data.iterrows():
                        new_class, created = StudentClass.objects.update_or_create(
                            name=row['name'],
                            defaults={'faculty': Faculty.objects.get(name=row['faculty'])}
                        )
                        
                    librarian = Account.objects.get(username=request.session['account'])
                    Log.objects.create(action=f"Import the data about classes from the file named '{filename}'",  account=librarian)
                    message = "File imported successfully"
                except Exception as e:
                    message = f"Error processing file: {e}"
                print(message)
                return render(request, "/view_classes")
        else:
            form = UploadFileForm()
        return render(request, "/view_classes")
    except Exception as e:
        return redirect('error_500')

def download_sample_class(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('error_404')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        
        columns = ["name", "faculty"]
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        faculties = Faculty.objects.all()
        faculty_names = [faculty.name for faculty in faculties]

        faculty_names_str = ','.join(faculty_names)
        dv = DataValidation(type="list", formula1=f'"{faculty_names_str}"', allow_blank=True)
        sheet.add_data_validation(dv)

        for row in range(2, 999):
            dv.add(sheet.cell(row=row, column=3))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample_class.xlsx'

        workbook.save(response)
        return response
    except Exception as e:
        return redirect('error_500')

def change_password(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_LIBRARIAN.value:
            return redirect('student_change_password')
        
        if request.method == "POST":
            old_password = request.POST['oldPassword']
            new_password = request.POST['newPassword']
            confirm_password = request.POST['confirmPassword']
            username = request.session['account']
            account = Account.objects.get(username=username)
            if account.password == old_password:
                if new_password == confirm_password:
                    account.password = new_password
                    account.save()
                    alert = True
                    librarian = Account.objects.get(username=request.session['account'])
                    Log.objects.create(action=f"Change password",  account=librarian)
                    return render(request, "change_password.html", {'alert':alert, 'account':account})
                else:
                    error_message = "Mật khẩu mới không khớp"
                    return render(request, "change_password.html", {'error_message':error_message, 'account':account})
            else:
                error_message = "Mật khẩu cũ không đúng"
                return render(request, "change_password.html", {'error_message':error_message, 'account':account})

        username = request.session['account']
        account = Account.objects.get(username=username)
        return render(request, "change_password.html", {'account':account})
    except Exception as e:
        return redirect('error_500')

def home(request):
    return render(request, 'library/index.html')

def login(request):
    try:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            try:
                account = Account.objects.get(username=username)
                if account.password == password:
                    request.session['account'] = account.username
                    request.session['role'] = account.role
                    if (account.role == RoleEnum.IS_STUDENT.value):
                        return redirect('student_view_loan')
                    elif (account.role == RoleEnum.IS_LIBRARIAN.value):
                        return redirect('view_loan_history')
                    elif (account.role == RoleEnum.IS_ADMIN.value):
                        return redirect('admin_view_logs')
                    else:
                        print(account.role)
                        print(RoleEnum.IS_STUDENT.value)
                        error_message = 'Tài khoản hoặc mật khẩu không đúng'
                        return render(request, 'library/login.html', {'error_message': error_message})
                else:
                    error_message = 'Tài khoản hoặc mật khẩu không đúng'
                    return render(request, 'library/login.html', {'error_message': error_message})
            except Account.DoesNotExist:
                error_message = 'Tài khoản hoặc mật khẩu không đúng'
                return render(request, 'library/login.html', {'error_message': error_message})
        else:
            return render(request, 'library/login.html')
    except Exception as e:
        print(e)
        return redirect('error_500')

def logout(request):
    try:
        if 'account' in request.session:
            del request.session['account']
            if 'role' in request.session:
                del request.session['role']
        
        return redirect('login')
    except Exception as e:
        return redirect('error_500')

# Student
def student_loan_history(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        
        student_ID = request.session['account']
        loan_list = Loan.objects.filter(student__id=student_id).prefetch_related('student')
        for loan in loan_list:
            loan.bookIds = ', '.join(str(book.id) for book in loan.books.all())

        return render(request, 'student_view_loan.html', {'loan_list': loan_list, 'today': date.today()})
    except Exception as e:
        return redirect('error_500')


def student_view_books(request, categoryId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_STUDENT.value:
            return redirect('error_404')
        
        if categoryId:
            books = Book.objects.filter(category=categoryId)
        else:
            books = Book.objects.all()
        categories = Category.objects.all()

        return render(request, "student_view_books.html", {'books':books, 'categories':categories})
    except Exception as e:
        return redirect('error_500')

def student_view_categories(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_STUDENT.value:
            return redirect('error_404')
        
        categories = Category.objects.annotate(quantity=Count('book'))
        return render(request, "student_view_categories.html", {'categories': categories})
    except Exception as e:
        return redirect('error_500')

def student_change_password(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_STUDENT.value:
            return redirect('error_404')
        
        if request.method == "POST":
            old_password = request.POST['oldPassword']
            new_password = request.POST['newPassword']
            confirm_password = request.POST['confirmPassword']
            username = request.session['account']
            account = Account.objects.get(username=username)
            if account.password == old_password:
                if new_password == confirm_password:
                    account.password = new_password
                    account.save()
                    alert = True
                    librarian = Account.objects.get(username=request.session['account'])
                    Log.objects.create(action=f"Change password",  account=librarian)
                    return render(request, "student_change_password.html", {'alert':alert, 'account':account})
                else:
                    error_message = "Mật khẩu mới không khớp"
                    return render(request, "student_change_password.html", {'error_message':error_message, 'account':account})
            else:
                error_message = "Mật khẩu cũ không đúng"
                return render(request, "student_change_password.html", {'error_message':error_message, 'account':account})

        username = request.session['account']
        account = Account.objects.get(username=username)
        return render(request, "student_change_password.html", {'account':account})
    except Exception as e:
        return redirect('error_500')

### Admin
# Account
def admin_view_loan_history(request, studentId=None):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        if studentId:
            loans = Loan.objects.filter(student=Student.objects.get(id=studentId))
        else:
            loans = Loan.objects.all()
        for loan in loans:
            loan.bookIds = ', '.join(str(book.id) for book in loan.books.all())

        return render(request, "admin_view_loan_history.html", {'loanList':loans})
    except Exception as e:
        return redirect('error_500')
    
def admin_view_accounts(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        accounts = Account.objects.exclude(role=RoleEnum.IS_ADMIN.value)
        return render(request, "admin_view_accounts.html", {'accounts':accounts})
    except Exception as e:
        print(e)
        return redirect('error_500')

def admin_upload_file_account(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        if request.method == 'POST' and request.FILES['file']:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                fs = FileSystemStorage()
                filename = fs.save(file.name, file)
                file_path = fs.path(filename)
                
                try:
                    data = pd.read_excel(file_path)
                    for index, row in data.iterrows():
                        account, created = Account.objects.update_or_create(
                            username=row['username'],
                            defaults={
                                'password': row['password'],
                                'role': RoleEnum.IS_LIBRARIAN.value,
                            }
                        )
                    alert = "File imported successfully"
                except Exception as e:
                    alert = f"Error processing file: {e}"

                return redirect("/admin_view_accounts")
        else:
            form = UploadFileForm()
        return redirect("/admin_view_accounts")
    except Exception as e:
        return redirect('error_500')

def admin_download_sample_account(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        columns = ["username", "password"]
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample_account.xlsx'

        workbook.save(response)
        return response
    except Exception as e:
        return redirect('error_500')
    
def admin_delete_account(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        account = Account.objects.get(id=id)
        account.delete()
        return redirect('admin_view_accounts')
    except Exception as e:
        return redirect('error_500')

def admin_reset_password_account(request, id):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        account = Account.objects.get(id=id)
        account.password = account.username
        account.save()
        return redirect('admin_view_accounts')
    except Exception as e:
        return redirect('error_500')
    
def admin_view_logs(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        logs = Log.objects.all()
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        return render(request, "admin_view_logs.html", {'logs':logs, 'months':months})
    except Exception as e:
        print(e)
        return redirect('error_500')
    
def admin_change_password(request):
    try:
        if 'account' not in request.session:
            return redirect('login')
        elif 'role' in request.session and request.session['role'] != RoleEnum.IS_ADMIN.value:
            return redirect('error_404')
        
        if request.method == "POST":
            old_password = request.POST['oldPassword']
            new_password = request.POST['newPassword']
            confirm_password = request.POST['confirmPassword']
            username = request.session['account']
            account = Account.objects.get(username=username)
            if account.password == old_password:
                if new_password == confirm_password:
                    account.password = new_password
                    account.save()
                    alert = True
                    return render(request, "admin_change_password.html", {'alert':alert, 'account':account})
                else:
                    error_message = "Mật khẩu mới không khớp"
                    return render(request, "admin_change_password.html", {'error_message':error_message, 'account':account})
            else:
                error_message = "Mật khẩu cũ không đúng"
                return render(request, "admin_change_password.html", {'error_message':error_message, 'account':account})

        username = request.session['account']
        account = Account.objects.get(username=username)
        return render(request, "admin_change_password.html", {'account':account})
    except Exception as e:
        print(e)
        return redirect('error_500')
    
def resetRole(request):
    # Account.objects.filter(username='102210130').update(role=RoleEnum.IS_STUDENT.value)

    # account = Account.objects.create(
    #     username='102210130',
    #     password='102210130',
    #     role=RoleEnum.IS_STUDENT.value
    # )
    # student = Student.objects.filter(id='102210130')[0]
    # student.account = account
    # student.save()
    for account in Account.objects.all():
        if Student.objects.filter(id=account.username).exists():
            student = Student.objects.get(id=account.username)
            student.account = account
            print(student)
            student.save()
    return redirect('view_books')